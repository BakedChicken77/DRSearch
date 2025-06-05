import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd

# Your existing imports and code...

def apply_excel_formatting(file_path, wrap_columns, freeze_row, table_style):
    # Open the created Excel file with openpyxl
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Wrap text in specified columns
    for col in wrap_columns:
        for cell in ws[col]:
            cell.alignment = Alignment(wrap_text=True)

    # Freeze the specified row
    ws.freeze_panes = ws[freeze_row]

    # Make all columns filterable
    table = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(
        name=table_style, 
        showFirstColumn=False,
        showLastColumn=False, 
        showRowStripes=True, 
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    # Save the changes
    wb.save(file_path)



# # Using this function

# # Call the formatting function
# apply_excel_formatting(
#     file_path=output_file_path,
#     wrap_columns=['D', 'F'],
#     freeze_row='A2',
#     table_style="TableStyleMedium9"
# )
